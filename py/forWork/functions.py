import pandas as pd
import numpy as np
import time
import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
from desktopmagic.screengrab_win32 import getRectAsImage
import pyautogui
import os

# 保存调用的各种函数

class function:
    def adjustFormat(fl,wraptext):
        # 调整单元格格式
        workbook = load_workbook(filename=fl)
        sheet = workbook.active
        s="QAZWSXEDCRFVTGBYHNUJMIOPKL"
        for item in s:
            colD = sheet[item]
            for c in colD:
        # max_rows = sheet.max_row  # 获取最大行
        # max_columns = sheet.max_column  # 获取最大列   
        # for i in range(1,max_rows+1):
        #     for j in range(1,max_columns+1):
        #         c=sheet.cell(i, j)
                if wraptext==1:
                     align=Alignment(horizontal='left',vertical='center',wrapText=True)
                else:
                    align=Alignment(horizontal='left',vertical='center')
                font = Font(name="微软雅黑",size=10)    
                border=Border(left=Side(style='thin'),  bottom=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'))
                
                c.font = font
                c.border = border
                c.alignment = align
        workbook.save(filename=fl)

    def toPercent(fl,columns):
        # 将对应行转换为百分数
        workbook = load_workbook(filename=fl)
        sheet = workbook.active
        for column in columns:
            # tmp=sheet['D']
            tmp=sheet[column]
            for i in tmp:
                if i.value!=2:
                    data= i.value
                    data = float('%.4f' % (float(data)))
                    i.value = data
                    i.number_format = '0.00%'

        workbook.save(filename=fl)
    
    def screenshot(fn):
        # 截图函数
        inMiddleXY=(4565,500)
        closePageXY=(3010,-1075)
        clickOkToLeaveXY=(4000,-850)
        basicPath=function.getmyCodeNexusPath("/py/forWork")
        pngPath=basicPath+"/png/"
        screenshotsXY=(2310,-600,4468,471)

        pyautogui.click(inMiddleXY)
        time.sleep(0.5)
        rect = getRectAsImage(screenshotsXY)
        rect.save(pngPath+fn+'.png', format='png')
        time.sleep(0.5)
        pyautogui.click(closePageXY)
        time.sleep(0.5)
        pyautogui.click(clickOkToLeaveXY)

    def getmyCodeNexusPath(fn=''):
        # 获得myCodeNexus的路径位置
        path=os.environ['myCodeNexusPath']+fn
        return path